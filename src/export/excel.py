from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ── Shared styles ─────────────────────────────────────────────────────────────
_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT  = Font(name="Calibri", bold=True,   color="FFFFFF", size=11)
_DATE_FONT    = Font(name="Calibri", size=10)
_CELL_FONT    = Font(name="Calibri", size=10)
_SECTION_FILL = PatternFill("solid", fgColor="2E5F8A")
_SECTION_FONT = Font(name="Calibri", bold=True,   color="FFFFFF", size=10)
_LABEL_FILL   = PatternFill("solid", fgColor="D6E4F0")
_LABEL_FONT   = Font(name="Calibri", bold=True,   color="1F3864", size=10)
_NOTE_FONT    = Font(name="Calibri", italic=True,  color="666666", size=9)
_THIN_BORDER  = Border(bottom=Side(style="thin", color="D9D9D9"))
_SECT_BORDER  = Border(
    top=Side(style="thin",   color="1F3864"),
    bottom=Side(style="thin", color="1F3864"),
)
_CENTER  = Alignment(horizontal="center", vertical="center")
_LEFT    = Alignment(horizontal="left",   vertical="center")
_PCT     = "0.00%"
_DEC4    = "0.0000"

_TICKER_DISPLAY = {"^GSPC": "S&P 500", "^DJI": "Dow Jones", "^IXIC": "NASDAQ"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _col(idx: int) -> str:
    return get_column_letter(idx)


def _disp(ticker: str) -> str:
    return _TICKER_DISPLAY.get(ticker, ticker)


def _max_drawdown(prices: pd.Series) -> float:
    peak = prices.cummax()
    return float(((prices - peak) / peak).min())


def _style_header_row(ws, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell           = ws.cell(row=1, column=c)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER


def _style_data_rows(ws, first_row: int, last_row: int, n_cols: int,
                     date_col: bool = True, num_format: str = "#,##0.00") -> None:
    for row in ws.iter_rows(min_row=first_row, max_row=last_row,
                            min_col=1, max_col=n_cols):
        for i, cell in enumerate(row):
            cell.border = _THIN_BORDER
            if date_col and i == 0:
                cell.font          = _DATE_FONT
                cell.alignment     = _CENTER
                cell.number_format = "YYYY-MM-DD"
            else:
                cell.font          = _CELL_FONT
                cell.alignment     = _CENTER
                cell.number_format = num_format


# ── Metrics + embedded Daily Returns in Adj Close Prices ──────────────────────

def _add_metrics_section(
    ws,
    closes: pd.DataFrame,
    n_data_rows: int,
    benchmark_ticker: str = None,
    risk_free_rate: float = 0.0,
) -> None:
    n_tickers  = len(closes.columns)
    last_price = n_data_rows + 1
    total_cols = n_tickers + 1

    # ── Embedded Daily Returns table ──────────────────────────────────────────
    # Layout: blank row, section header, column header, then n_data_rows-1 data rows
    dr_hdr     = last_price + 2
    dr_col_hdr = dr_hdr + 1
    dr_first   = dr_col_hdr + 1
    dr_last    = dr_first + n_data_rows - 2   # n_data_rows - 1 return rows

    for c in range(1, total_cols + 1):
        cell           = ws.cell(row=dr_hdr, column=c)
        cell.fill      = _SECTION_FILL
        cell.font      = _SECTION_FONT
        cell.alignment = _CENTER
        cell.border    = _SECT_BORDER
    ws.cell(row=dr_hdr, column=1).value = "DAILY RETURNS"

    ws.cell(row=dr_col_hdr, column=1).value = "Date"
    for j, ticker in enumerate(closes.columns):
        ws.cell(row=dr_col_hdr, column=j + 2).value = ticker
    for c in range(1, total_cols + 1):
        cell           = ws.cell(row=dr_col_hdr, column=c)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER

    for i in range(1, n_data_rows):
        ret_row   = dr_first + i - 1
        price_row = i + 1
        next_row  = i + 2

        date_cell               = ws.cell(row=ret_row, column=1)
        date_cell.value         = f"=A{next_row}"
        date_cell.font          = _DATE_FONT
        date_cell.alignment     = _CENTER
        date_cell.number_format = "YYYY-MM-DD"
        date_cell.border        = _THIN_BORDER

        for j in range(n_tickers):
            c    = _col(j + 2)
            cell = ws.cell(row=ret_row, column=j + 2)
            cell.value         = f"={c}{next_row}/{c}{price_row}-1"
            cell.font          = _CELL_FONT
            cell.alignment     = _CENTER
            cell.number_format = _PCT
            cell.border        = _THIN_BORDER

    # ── Metrics start after a blank row ───────────────────────────────────────
    sec_start = dr_last + 2

    ticker_list   = list(closes.columns)
    bench_col_idx = (ticker_list.index(benchmark_ticker) + 2
                     if benchmark_ticker and benchmark_ticker in ticker_list
                     else None)
    bench_col = _col(bench_col_idx) if bench_col_idx else None

    # ── Row builders ──────────────────────────────────────────────────────────
    def section_header(row_num: int, title: str) -> None:
        for c in range(1, total_cols + 1):
            cell           = ws.cell(row=row_num, column=c)
            cell.fill      = _SECTION_FILL
            cell.font      = _SECTION_FONT
            cell.alignment = _CENTER
            cell.border    = _SECT_BORDER
        ws.cell(row=row_num, column=1).value = title

    def metric_row(row_num: int, label: str, values: list,
                   num_format: str = _PCT) -> None:
        lc           = ws.cell(row=row_num, column=1)
        lc.value     = label
        lc.fill      = _LABEL_FILL
        lc.font      = _LABEL_FONT
        lc.alignment = _LEFT
        lc.border    = _THIN_BORDER
        for i, val in enumerate(values):
            cell           = ws.cell(row=row_num, column=i + 2)
            cell.font      = _CELL_FONT
            cell.alignment = _CENTER
            cell.border    = _THIN_BORDER
            if val == "—":
                cell.value         = "—"
                cell.number_format = "@"
            else:
                cell.value         = val
                cell.number_format = num_format

    def cell_rng(col_idx: int) -> str:
        """Same-sheet range for the embedded daily returns column."""
        c = _col(col_idx)
        return f"{c}{dr_first}:{c}{dr_last}"

    # ── Return Metrics ────────────────────────────────────────────────────────
    section_header(sec_start, "RETURN METRICS")

    metric_row(sec_start + 1, "Total Return",
        [f"=({_col(j+2)}{last_price}-{_col(j+2)}2)/{_col(j+2)}2"
         for j in range(n_tickers)])

    metric_row(sec_start + 2, "CAGR",
        [f"=({_col(j+2)}{last_price}/{_col(j+2)}2)"
         f"^(252/(ROWS({_col(j+2)}2:{_col(j+2)}{last_price})-1))-1"
         for j in range(n_tickers)])

    metric_row(sec_start + 3, "Annualized Return",
        [f"=AVERAGE({cell_rng(j+2)})*252"
         for j in range(n_tickers)])

    metric_row(sec_start + 4, "Annualized Volatility",
        [f"=STDEV({cell_rng(j+2)})*SQRT(252)"
         for j in range(n_tickers)])

    # ── Risk Metrics ──────────────────────────────────────────────────────────
    section_header(sec_start + 6, "RISK METRICS  (95% confidence)")

    metric_row(sec_start + 7, "Max Drawdown *",
        [_max_drawdown(closes.iloc[:, j]) for j in range(n_tickers)])

    metric_row(sec_start + 8, "VaR 95% Daily",
        [f"=-PERCENTILE({cell_rng(j+2)},0.05)"
         for j in range(n_tickers)])

    metric_row(sec_start + 9, "ES 95% Daily",
        [f"=-AVERAGEIF({cell_rng(j+2)},\"<=\"&PERCENTILE({cell_rng(j+2)},0.05))"
         for j in range(n_tickers)])

    note = ws.cell(row=sec_start + 11, column=1)
    note.value = ("* Max Drawdown: pre-computed (max peak-to-trough decline). "
                  "All other metrics use Excel formulas above.")
    note.font  = _NOTE_FONT

    # ── Risk-Adjusted Ratios ──────────────────────────────────────────────────
    drf = risk_free_rate / 252

    section_header(sec_start + 13, "RISK-ADJUSTED RATIOS")

    metric_row(sec_start + 14, "Sharpe Ratio",
        [f"=(AVERAGE({cell_rng(j+2)})*252-{risk_free_rate})"
         f"/(STDEV({cell_rng(j+2)})*SQRT(252))"
         for j in range(n_tickers)],
        num_format=_DEC4)

    metric_row(sec_start + 15, "Sortino Ratio",
        [f"=(AVERAGE({cell_rng(j+2)})*252-{risk_free_rate})"
         f"/(SQRT(SUMPRODUCT((({cell_rng(j+2)}-{drf})*({cell_rng(j+2)}<{drf}))^2)"
         f"/ROWS({cell_rng(j+2)}))*SQRT(252))"
         for j in range(n_tickers)],
        num_format=_DEC4)

    # ── Relative Metrics ──────────────────────────────────────────────────────
    if bench_col is None:
        ws.column_dimensions[_col(1)].width = 26
        return

    bench_rng_str    = f"{bench_col}{dr_first}:{bench_col}{dr_last}"
    rel_start        = sec_start + 18
    beta_row         = rel_start + 1
    excess_ret_row   = rel_start + 3
    tracking_err_row = rel_start + 4
    bench_label      = _disp(benchmark_ticker)

    section_header(rel_start, f"RELATIVE METRICS vs {bench_label}")

    def rel_vals(formula_fn, num_format=_PCT):
        vals = []
        for j in range(n_tickers):
            col_idx = j + 2
            if col_idx == bench_col_idx:
                vals.append("—")
            else:
                c = _col(col_idx)
                r = f"{c}{dr_first}:{c}{dr_last}"
                vals.append(formula_fn(c, r))
        return vals

    # Treynor in the Risk-Adjusted section above (sec_start+16).
    # Benchmark β vs itself = 1 by definition, so divide by 1 (= equity risk premium).
    def treynor_vals():
        vals = []
        for j in range(n_tickers):
            col_idx = j + 2
            c       = _col(col_idx)
            r       = f"{c}{dr_first}:{c}{dr_last}"
            divisor = "1" if col_idx == bench_col_idx else f"{c}{beta_row}"
            vals.append(f"=(AVERAGE({r})*252-{risk_free_rate})/{divisor}")
        return vals

    metric_row(sec_start + 16, "Treynor Ratio", treynor_vals(), num_format=_DEC4)

    metric_row(beta_row, "Beta",
        rel_vals(lambda c, r: f"=SLOPE({r},{bench_rng_str})"),
        num_format=_DEC4)

    metric_row(rel_start + 2, "Alpha (Annualized)",
        rel_vals(lambda c, r:
            f"=(AVERAGE({r})*252-{risk_free_rate})"
            f"-{c}{beta_row}*(AVERAGE({bench_rng_str})*252-{risk_free_rate})"))

    metric_row(excess_ret_row, "Excess Return",
        rel_vals(lambda c, r:
            f"=AVERAGE({r})*252-AVERAGE({bench_rng_str})*252"))

    metric_row(tracking_err_row, "Tracking Error",
        rel_vals(lambda c, r:
            f"=SQRT(SUMPRODUCT(({r}-{bench_rng_str}"
            f"-AVERAGE({r})+AVERAGE({bench_rng_str}))^2)"
            f"/(ROWS({r})-1))*SQRT(252)"))

    metric_row(rel_start + 5, "Information Ratio",
        rel_vals(lambda c, r:
            f"={c}{excess_ret_row}/{c}{tracking_err_row}"),
        num_format=_DEC4)

    ws.column_dimensions[_col(1)].width = 26


# ── Portfolio (Equal Weight) sheet ───────────────────────────────────────────

def _build_portfolio_sheet(
    wb,
    closes: pd.DataFrame,
    n_data_rows: int,
    dr_first: int,
    benchmark_ticker: str = None,
    risk_free_rate: float = 0.0,
    weights: dict = None,
) -> None:
    """
    Add a 'Portfolio (Equal Weight)' sheet containing:
      - weights table
      - daily portfolio returns column (weighted sum of individual returns)
      - cumulative return column
      - all metric formula sections
    Daily returns are sourced from the embedded table in 'Adj Close Prices'.
    """
    ticker_list  = list(closes.columns)
    port_tickers = [t for t in ticker_list if t != benchmark_ticker]
    if not port_tickers:
        return

    n_port    = len(port_tickers)
    n_returns = n_data_rows - 1
    dr_last   = dr_first + n_returns - 1

    # Resolve weights: use provided weights or default to equal weight
    if weights and all(t in weights for t in port_tickers):
        port_weights = {t: weights[t] for t in port_tickers}
        is_equal     = len(set(round(w, 6) for w in port_weights.values())) == 1
        sheet_title  = "Portfolio (Equal Weight)" if is_equal else "Portfolio"
    else:
        eq = 1.0 / n_port
        port_weights = {t: eq for t in port_tickers}
        is_equal     = True
        sheet_title  = "Portfolio (Equal Weight)"

    # Benchmark column in Adj Close Prices (same indices as embedded daily returns)
    bench_col_idx = (ticker_list.index(benchmark_ticker) + 2
                     if benchmark_ticker and benchmark_ticker in ticker_list
                     else None)
    bench_col = _col(bench_col_idx) if bench_col_idx else None
    bench_rng = (f"'Adj Close Prices'!{bench_col}{dr_first}:'Adj Close Prices'!{bench_col}{dr_last}"
                 if bench_col else None)

    # Portfolio ticker column letters in Adj Close Prices sheet
    port_cols = [_col(ticker_list.index(t) + 2) for t in port_tickers]

    # Pre-compute max drawdown using actual weights
    ret_matrix  = closes[port_tickers].pct_change().dropna()
    w_series    = pd.Series(port_weights)
    port_ret_py = ret_matrix.mul(w_series, axis=1).sum(axis=1)
    cum         = (1 + port_ret_py).cumprod()
    max_dd      = float(((cum - cum.cummax()) / cum.cummax()).min())

    ws = wb.create_sheet(sheet_title)

    # ── Row layout ────────────────────────────────────────────────────────────
    WT_HDR   = 4
    WT_COL   = 5
    WT_START = 6
    WT_END   = WT_START + n_port - 1

    RET_HDR   = WT_END + 2
    RET_COL   = RET_HDR + 1
    RET_FIRST = RET_COL + 1
    RET_LAST  = RET_FIRST + n_returns - 1

    MET  = RET_LAST + 2
    drf  = risk_free_rate / 252
    rf   = risk_free_rate

    port_rng = f"B{RET_FIRST}:B{RET_LAST}"

    # ── Inner helpers ─────────────────────────────────────────────────────────
    def sec_hdr(row, title, n_cols=2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = _SECTION_FILL; cell.font = _SECTION_FONT
            cell.alignment = _CENTER;  cell.border = _SECT_BORDER
        ws.cell(row=row, column=1).value = title

    def col_hdr(row, labels):
        for i, lbl in enumerate(labels):
            cell = ws.cell(row=row, column=i + 1)
            cell.value = lbl; cell.fill = _HEADER_FILL
            cell.font  = _HEADER_FONT; cell.alignment = _CENTER

    def kv(row, label, value, num_fmt=_PCT):
        a = ws.cell(row=row, column=1)
        a.value = label; a.fill = _LABEL_FILL; a.font = _LABEL_FONT
        a.alignment = _LEFT; a.border = _THIN_BORDER
        b = ws.cell(row=row, column=2)
        b.value = value; b.font = _CELL_FONT
        b.alignment = _CENTER; b.number_format = num_fmt; b.border = _THIN_BORDER

    # ── Title ─────────────────────────────────────────────────────────────────
    t = ws.cell(row=1, column=1)
    t.value = "EQUAL-WEIGHT PORTFOLIO ANALYSIS"
    t.font  = Font(name="Calibri", bold=True, size=14, color="1F3864")
    t.alignment = _LEFT

    bench_label = _disp(benchmark_ticker) if benchmark_ticker else "None"
    s = ws.cell(row=2, column=1)
    s.value = (f"Tickers: {', '.join(port_tickers)}  |  "
               f"Benchmark: {bench_label}  |  "
               f"Risk-Free Rate: {rf:.2%}")
    s.font  = Font(name="Calibri", size=10, color="444444")

    # ── Weights table ─────────────────────────────────────────────────────────
    sec_hdr(WT_HDR, "PORTFOLIO WEIGHTS")
    col_hdr(WT_COL, ["Ticker", "Weight"])
    for i, ticker in enumerate(port_tickers):
        row = WT_START + i
        a = ws.cell(row=row, column=1)
        a.value = ticker; a.font = _CELL_FONT; a.alignment = _CENTER; a.border = _THIN_BORDER
        b = ws.cell(row=row, column=2)
        b.value = port_weights[ticker]; b.font = _CELL_FONT; b.alignment = _CENTER
        b.number_format = _PCT; b.border = _THIN_BORDER

    # ── Portfolio daily returns ───────────────────────────────────────────────
    sec_hdr(RET_HDR, "PORTFOLIO DAILY RETURNS", n_cols=3)
    col_hdr(RET_COL, ["Date", "Daily Return", "Cum. Return"])

    def weighted_return(dr_row: int) -> str:
        terms = [f"{port_weights[t]:.8f}*'Adj Close Prices'!{c}{dr_row}"
                 for t, c in zip(port_tickers, port_cols)]
        return "=" + "+".join(terms)

    for i in range(n_returns):
        r      = RET_FIRST + i
        dr_row = dr_first + i

        d = ws.cell(row=r, column=1)
        d.value = f"='Adj Close Prices'!A{dr_row}"
        d.font = _DATE_FONT; d.alignment = _CENTER
        d.number_format = "YYYY-MM-DD"; d.border = _THIN_BORDER

        b = ws.cell(row=r, column=2)
        b.value = weighted_return(dr_row)
        b.font = _CELL_FONT; b.alignment = _CENTER
        b.number_format = _PCT; b.border = _THIN_BORDER

        c = ws.cell(row=r, column=3)
        c.value = f"=1+B{r}" if r == RET_FIRST else f"=C{r-1}*(1+B{r})"
        c.font = _CELL_FONT; c.alignment = _CENTER
        c.number_format = "#,##0.0000"; c.border = _THIN_BORDER

    # ── Return Metrics ────────────────────────────────────────────────────────
    sec_hdr(MET, "RETURN METRICS")
    kv(MET+1, "Total Return",           f"=C{RET_LAST}-1")
    kv(MET+2, "CAGR",
       f"=C{RET_LAST}^(252/ROWS({port_rng}))-1")
    kv(MET+3, "Annualized Return",      f"=AVERAGE({port_rng})*252")
    kv(MET+4, "Annualized Volatility",  f"=STDEV({port_rng})*SQRT(252)")

    # ── Risk Metrics ──────────────────────────────────────────────────────────
    sec_hdr(MET+6, "RISK METRICS  (95% confidence)")
    kv(MET+7, "Max Drawdown *",  max_dd)
    kv(MET+8, "VaR 95% Daily",   f"=-PERCENTILE({port_rng},0.05)")
    kv(MET+9, "ES 95% Daily",
       f"=-AVERAGEIF({port_rng},\"<=\"&PERCENTILE({port_rng},0.05))")

    # ── Risk-Adjusted Ratios ──────────────────────────────────────────────────
    sec_hdr(MET+11, "RISK-ADJUSTED RATIOS")
    kv(MET+12, "Sharpe Ratio",
       f"=(AVERAGE({port_rng})*252-{rf})/(STDEV({port_rng})*SQRT(252))",
       num_fmt=_DEC4)
    kv(MET+13, "Sortino Ratio",
       f"=(AVERAGE({port_rng})*252-{rf})"
       f"/(SQRT(SUMPRODUCT(((({port_rng})-{drf})*(({port_rng})<{drf}))^2)"
       f"/ROWS({port_rng}))*SQRT(252))",
       num_fmt=_DEC4)

    # ── Relative Metrics ──────────────────────────────────────────────────────
    if bench_rng:
        beta_row = MET + 17
        er_row   = MET + 19
        te_row   = MET + 20
        kv(MET+14, "Treynor Ratio",
           f"=(AVERAGE({port_rng})*252-{rf})/B{beta_row}", num_fmt=_DEC4)
        sec_hdr(MET+16, f"RELATIVE METRICS vs {bench_label}")
        kv(beta_row, "Beta",
           f"=SLOPE({port_rng},{bench_rng})", num_fmt=_DEC4)
        kv(MET+18, "Alpha (Annualized)",
           f"=(AVERAGE({port_rng})*252-{rf})"
           f"-B{beta_row}*(AVERAGE({bench_rng})*252-{rf})")
        kv(er_row,  "Excess Return",
           f"=AVERAGE({port_rng})*252-AVERAGE({bench_rng})*252")
        kv(te_row,  "Tracking Error",
           f"=SQRT(SUMPRODUCT(({port_rng}-{bench_rng}"
           f"-AVERAGE({port_rng})+AVERAGE({bench_rng}))^2)"
           f"/(ROWS({port_rng})-1))*SQRT(252)")
        kv(MET+21, "Information Ratio",
           f"=B{er_row}/B{te_row}", num_fmt=_DEC4)

    # Note
    note = ws.cell(row=MET + 23, column=1)
    note.value = ("* Max Drawdown: pre-computed (max peak-to-trough decline). "
                  "All other metrics use Excel formulas.")
    note.font  = _NOTE_FONT

    # ── Growth of $10,000 chart (anchored at E1) ─────────────────────────────
    from src.export.charts import growth_of_10k
    bench_ret_py = (closes[benchmark_ticker].pct_change().dropna()
                    if benchmark_ticker and benchmark_ticker in closes.columns
                    else None)
    chart_buf = growth_of_10k(
        portfolio_returns=port_ret_py,
        benchmark_returns=bench_ret_py,
        portfolio_label="Portfolio",
        benchmark_label=benchmark_ticker or "Benchmark",
        width_in=8.0,
        height_in=3.2,
    )
    xl_img        = XLImage(chart_buf)
    xl_img.width  = 576
    xl_img.height = 230
    ws.add_image(xl_img, "E1")

    # ── Column widths & formatting ────────────────────────────────────────────
    ws.column_dimensions[_col(1)].width = 28
    ws.column_dimensions[_col(2)].width = 18
    ws.column_dimensions[_col(3)].width = 16
    ws.freeze_panes             = "A2"
    ws.sheet_view.showGridLines = False


# ── Main export ───────────────────────────────────────────────────────────────

def export_pricing(
    price_data: dict,
    output_dir: Path,
    start_date: str,
    end_date: str,
    benchmark_ticker: str = None,
    risk_free_rate: float = 0.0,
    weights: dict = None,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    closes = pd.DataFrame(
        {_disp(ticker): df["adj_close"] for ticker, df in price_data.items()}
    ).sort_index()

    filename = f"Pricing_Data_{start_date}_to_{end_date}.xlsx"
    filepath = output_dir / filename

    n_data_rows   = len(closes)
    n_ticker_cols = len(closes.columns)
    n_cols        = n_ticker_cols + 1

    # Row where the embedded daily returns data starts in Adj Close Prices:
    #   last_price (n_data_rows+1) + blank + section_hdr + col_hdr + 1 = n_data_rows + 5
    dr_first = n_data_rows + 5

    with pd.ExcelWriter(filepath, engine="openpyxl", date_format="YYYY-MM-DD") as writer:
        closes.to_excel(writer, sheet_name="Adj Close Prices", index=True)

        ws = writer.sheets["Adj Close Prices"]
        wb = writer.book

        _style_header_row(ws, n_cols)
        ws.cell(row=1, column=1).value = "Date"
        _style_data_rows(ws, 2, n_data_rows + 1, n_cols)

        bench_display = _disp(benchmark_ticker) if benchmark_ticker else None
        _add_metrics_section(ws, closes, n_data_rows, bench_display, risk_free_rate)
        _build_portfolio_sheet(wb, closes, n_data_rows, dr_first,
                               bench_display, risk_free_rate, weights)

        ws.column_dimensions[_col(1)].width = 26
        for ci in range(2, n_cols + 1):
            ws.column_dimensions[_col(ci)].width = 18

        ws.freeze_panes             = "A2"
        ws.sheet_view.showGridLines = False

    return filepath
